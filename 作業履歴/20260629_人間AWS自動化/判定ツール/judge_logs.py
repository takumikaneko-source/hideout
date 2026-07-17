# -*- coding: utf-8 -*-
"""
アラートログ判定ツール（Phase3）

ログ収集ツールが転記したエラーログ（D列）を読み取り、
「対応要否」と「影響度」を判定して G列以降に書き込みます。

判定の流れ（1行ごと）:
  1. rules.yaml と機械照合（AI不使用・費用ゼロ）→ 一致すれば即断
  2. cases.jsonl から類似の過去事例を検索（機械処理）
  3. AI（Claude）が guidelines.md / system_context.md / 類似事例を
     もとに判定
  4. 確信度が低い場合のみ: 3観点（業務影響/復旧・再処理/事例照合）で
     再判定し多数決 → それでも割れたら上位モデルで最終判定

安全のための決まりごと:
  - AWS には一切アクセスしません（読み取りも含めて何もしない）
  - APIキーは環境変数 ANTHROPIC_API_KEY からのみ読みます（ソース直書き禁止）
  - 入力ファイルは変更しません（コピーを作って書き込みます）
  - E列（運用チームの対応要不要）には触りません。最終判断は運用チームです
"""

import io
import json
import os
import re
import shutil
import subprocess
import sys
import datetime

# 社内ネットワークの TLS 検査プロキシ対応（OSの証明書ストアを使う）
try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass  # 社外環境など truststore が無くても動くようにする

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import yaml

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---- 設定読み込み（settings.py が無くても既定値で動く） ----
try:
    import settings as _s
except ImportError:
    _s = object()

def _get(name, default):
    return getattr(_s, name, default)

def _path(p):
    """settings の相対パスをこのファイルのフォルダ基準で絶対パス化する"""
    if p is None:
        return None
    return p if os.path.isabs(p) else os.path.normpath(os.path.join(BASE_DIR, p))

INPUT_XLSX      = _path(_get("INPUT_XLSX", r"..\配布_ログ収集ツール\OUTPUT\slack通知アラート収集_2026年度_エラーログ収集後.xlsx"))
OUTPUT_DIR      = _path(_get("OUTPUT_DIR", None))
OUTPUT_XLSX     = _get("OUTPUT_XLSX", None)
SHEET_NAME      = _get("SHEET_NAME", None)
KNOWLEDGE_DIR   = _path(_get("KNOWLEDGE_DIR", r"..\ナレッジ外だし"))
BACKEND         = _get("BACKEND", "api")
CLAUDE_CMD      = _get("CLAUDE_CMD", "claude")
CLI_TIMEOUT     = _get("CLI_TIMEOUT", 120)
JUDGE_MODE      = _get("JUDGE_MODE", "pipeline")
MODEL_WORKER    = _get("MODEL_WORKER", "claude-haiku-4-5")
MODEL_ORCHESTRATOR = _get("MODEL_ORCHESTRATOR", "claude-opus-4-8")
MODEL_JUDGE     = _get("MODEL_JUDGE", "claude-sonnet-5")
MODEL_ESCALATION = _get("MODEL_ESCALATION", "claude-opus-4-8")
MULTI_PERSONA   = _get("MULTI_PERSONA", True)
ESCALATE_ON_LOW_CONFIDENCE = _get("ESCALATE_ON_LOW_CONFIDENCE", True)
TOP_N_CASES     = _get("TOP_N_CASES", 5)
MAX_LOG_CHARS   = _get("MAX_LOG_CHARS", 4000)
MAX_TOKENS      = _get("MAX_TOKENS", 8000)
DRY_RUN         = _get("DRY_RUN", False)
MAX_ROWS        = _get("MAX_ROWS", 200)
RE_JUDGE        = _get("RE_JUDGE", False)
# 判定差分の記録（別ファイル・永続）。判定済みエクセルは毎回洗い替えされるため、
# AI判定と運用判定の差分はここに蓄積する。
DIFF_RECORD      = _get("DIFF_RECORD", True)
DIFF_RECORD_XLSX = _path(_get("DIFF_RECORD_XLSX", r".\OUTPUT\AI判定_運用差分記録.xlsx"))

# 列の割り当て（A〜F はログ収集ツールと共通。G 以降が本ツールの書き込み先）
COL_DATE, COL_ALARM, COL_TIME, COL_ERRORS, COL_TAIOU, COL_LINK = 1, 2, 3, 4, 5, 6
COL_AI_ACTION, COL_AI_IMPACT, COL_AI_REASON, COL_AI_CONF, COL_AI_METHOD = 7, 8, 9, 10, 11
HEADER_ROW = 1

SERVICES = {"aegis", "blitz", "buster", "duel", "strike"}

# AI の回答形式（構造化出力）。この形の JSON だけが返ることが API 側で保証される
JUDGE_SCHEMA = {
    "type": "object",
    "properties": {
        "action":     {"type": "string", "enum": ["要", "静観", "不要"],
                       "description": "対応要否"},
        "impact":     {"type": "string", "enum": ["大", "中", "小"],
                       "description": "影響度"},
        "confidence": {"type": "string", "enum": ["高", "中", "低"],
                       "description": "判定の確信度。材料不足なら低"},
        "reason":     {"type": "string",
                       "description": "判定理由。【結論】【根拠】【確認事項】の3段落を、"
                                      "各段落の間を改行(\\n)で区切って書く。各段落2〜3文。"},
        "flow_note":  {"type": "string",
                       "description": "上流起因の可能性・下流への波及の所見。無ければ「特になし」"},
    },
    "required": ["action", "impact", "confidence", "reason", "flow_note"],
    "additionalProperties": False,
}


# ============================================================
# ナレッジの読み込み
# ============================================================

def load_knowledge():
    """ナレッジ4ファイルを読み込む。無いファイルがあれば止める"""
    def read_text(name):
        p = os.path.join(KNOWLEDGE_DIR, name)
        if not os.path.exists(p):
            sys.exit(f"[エラー] ナレッジファイルが見つかりません: {p}")
        return io.open(p, encoding="utf-8").read()

    rules = yaml.safe_load(read_text("rules.yaml"))
    guidelines = read_text("guidelines.md")
    system_context = read_text("system_context.md")

    cases = []
    for line in read_text("cases.jsonl").splitlines():
        line = line.strip()
        if line:
            try:
                cases.append(json.loads(line))
            except json.JSONDecodeError:
                pass  # 壊れた行は飛ばす（件数は多いので1行欠けても影響小）
    return rules, cases, guidelines, system_context


# ============================================================
# アラーム名からの情報抽出（build_cases.py と同じ考え方）
# ============================================================

ENV_RE = re.compile(r"-(prd|stgtmp|stg|dev\d?|dev)-")

def alarm_context(alarm_name, log_text):
    """アラーム名とログから service / env / source_type を推定する"""
    alarm = (alarm_name or "").strip()
    head = alarm.split("-", 1)[0]
    service = head if head in SERVICES else ""
    m = ENV_RE.search(alarm)
    env = m.group(1) if m else ""
    blob = (alarm + " " + (log_text or "")[:500]).lower()
    if any(x in blob for x in ("sfn", "statemachine", "states", "stepfunction", "ステートマシン")):
        source_type = "stepfunctions"
    elif "ecs" in blob or "fargate" in blob:
        source_type = "ecs"
    else:
        source_type = "lambda"
    return {"alarm_name": alarm, "service": service, "env": env,
            "source_type": source_type, "log_text": log_text or ""}


# ============================================================
# 1. rules.yaml による機械判定（AI不使用）
# ============================================================

def match_rule_when(when, ctx):
    """ルールの when 条件をすべて満たすか（AND）を判定する"""
    log = ctx["log_text"]
    for key, val in when.items():
        if key == "alarm_name_prefix":
            if not ctx["alarm_name"].startswith(val):
                return False
        elif key == "alarm_name_equals":
            if ctx["alarm_name"] != val:
                return False
        elif key in ("service", "env", "source_type", "namespace"):
            if ctx.get(key, "") != val:
                return False
        elif key == "log_contains_any":
            if not any(s in log for s in val):
                return False
        elif key == "log_contains_all":
            if not all(s in log for s in val):
                return False
        elif key == "log_regex":
            if not re.search(val, log):
                return False
        else:
            return False  # 未知の条件キーは安全側（不一致）に倒す
    return True


def apply_rules(rules, ctx):
    """上から順に照合し、最初に一致したルールを返す（無ければ None）"""
    for rule in rules.get("rules", []):
        if match_rule_when(rule.get("when", {}), ctx):
            return rule
    return None


# ============================================================
# 2. cases.jsonl からの類似事例検索（機械処理）
# ============================================================

KEYWORD_RE = re.compile(r"[A-Za-z][A-Za-z0-9_]{3,}")

# ほぼ全ログに含まれ判別に役立たない語（スコアの水増しを防ぐ）
_STOPWORDS = {"ERROR", "Error", "error", "WARNING", "Warning", "INFO",
              "Exception", "Message", "None", "null", "true", "false",
              "lambda", "function", "aws"}

def find_similar_cases(ctx, cases, top_n):
    """アラーム名の一致とログ中のキーワード重なりで類似事例を採点して上位を返す"""
    log = ctx["log_text"]
    scored = []
    for c in cases:
        score = 0
        c_alarm = c.get("alarm_name") or ""
        if c_alarm and c_alarm == ctx["alarm_name"]:
            score += 100  # 同一アラームは最重視
        elif c_alarm and ctx["service"] and c_alarm.startswith(ctx["service"]):
            score += 3
        if c.get("service") and c["service"] == ctx["service"]:
            score += 5
        # 過去事例のエラー要約に含まれる英字キーワードが今回のログにもあるか
        for kw in set(KEYWORD_RE.findall(c.get("error_summary") or "")) - _STOPWORDS:
            if kw in log:
                score += 3
        if not c.get("needs_review"):
            score += 2  # 判定理由が明確な事例を優先
        # 集約版では count（過去の発生回数）が多いほど確立したパターン＝わずかに優先
        cnt = c.get("count") or 0
        if cnt >= 100:
            score += 3
        elif cnt >= 10:
            score += 2
        elif cnt >= 3:
            score += 1
        if score > 0:
            scored.append((score, c))
    scored.sort(key=lambda x: -x[0])
    return [c for _, c in scored[:top_n]]


def format_cases(similar):
    """類似事例を AI に渡すテキストへ整形する"""
    if not similar:
        return "（類似事例なし）"
    lines = []
    for i, c in enumerate(similar, 1):
        # 集約版なら過去の発生回数を添える（頻出＝よくあるパターンの手掛かり）
        freq = f"（過去{c['count']}回）" if c.get("count") else ""
        lines.append(
            f"{i}. [{c.get('issue_key','')}] alarm={c.get('alarm_name','')}{freq} "
            f"暫定判定={c.get('action_hint','')}\n"
            f"   エラー: {(c.get('error_summary') or '')[:150]}\n"
            f"   判定理由: {(c.get('reason') or '')[:200]}"
        )
    return "\n".join(lines)


# ============================================================
# 3. AI 判定（バックエンド: API / Claude Code サブスク）
# ============================================================

ROLE_TEXT = (
    "あなたは taxi-cloud（タクシー動態管理システム）の運用アラート判定を支援するAIです。"
    "CloudWatchアラームで検知されたエラーログについて、対応要否（要/静観/不要）と"
    "影響度（大/中/小）を判定します。\n"
    "- 判定基準とシステム背景（処理フローと影響伝播）に必ず照らして判断すること\n"
    "- エラー箇所の前後関係（上流起因のサインが無いか、下流へ波及するか、"
    "次サイクルで自然回復するか）を必ず考慮すること\n"
    "- 過去の類似事例があればそれに倣うこと。ただし暫定判定(action_hint)は機械推定なので"
    "理由の文面を重視すること\n"
    "- 断定できない場合は確信度を「低」とし、安全側（対応要）に倒すこと\n"
    "- 最終判断は運用チームの目視で行うため、判断根拠を簡潔かつ明確に述べること\n"
    "- reason は次の3段落を、各段落を必ず改行(\\n)で区切り、行頭にラベルを付けて書くこと。"
    "各段落は2〜3文程度に収め、冗長にしないこと:\n"
    "  【結論】なぜその対応要否・影響度なのかを一言で\n"
    "  【根拠】判定基準・過去事例・ログの事実など拠り所を簡潔に\n"
    "  【確認事項】運用チームが目視で確かめるべき点（無ければ「特になし」）"
)


# ---- LLM 呼び出し層（バックエンドで API / Claude Code を切り替える）----

_client = None  # API バックエンド用（main で生成）

# 判定エラーをバックエンド非依存で扱うための例外
class RateLimitHit(Exception): pass
class AuthError(Exception): pass
class ConnError(Exception): pass


class _Usage:
    """API/CLI 双方で使う使用量の入れ物（無い項目は0）"""
    def __init__(self, i=0, o=0, cr=0, cw=0, cost=None):
        self.input_tokens = i
        self.output_tokens = o
        self.cache_read_input_tokens = cr
        self.cache_creation_input_tokens = cw
        self.cost = cost


# API のモデルID → Claude Code の --model 別名
CLI_MODEL = {
    "claude-opus-4-8": "opus", "claude-opus-4-7": "opus",
    "claude-sonnet-5": "sonnet", "claude-sonnet-4-6": "sonnet",
    "claude-haiku-4-5": "haiku",
}
def _cli_model(model):
    return CLI_MODEL.get(model, model)


_JSON_RE = re.compile(r"\{.*\}", re.S)
def _parse_json(text):
    """テキストから最初のJSONオブジェクトを取り出して辞書にする（前後の説明やコードフェンスに耐える）"""
    m = _JSON_RE.search(text or "")
    if not m:
        raise RuntimeError("JSON応答を解釈できませんでした: " + (text or "")[:120])
    return json.loads(m.group(0))


def build_pipeline_system(guidelines, system_context):
    """案A用: 役割＋判定基準＋システム背景をまとめた system 文字列（毎回同じ＝キャッシュが効く）"""
    return (ROLE_TEXT + "\n\n# 判定基準（guidelines）\n" + guidelines +
            "\n\n# システム背景（system_context）\n" + system_context)


def _api_call(model, system_text, user_text, schema):
    """Anthropic API を1回呼ぶ。system は安定文字列なのでキャッシュ指定を付ける"""
    import anthropic
    system = [{"type": "text", "text": system_text, "cache_control": {"type": "ephemeral"}}]
    kwargs = dict(model=model, max_tokens=MAX_TOKENS, system=system,
                  messages=[{"role": "user", "content": user_text}])
    if schema:
        kwargs["output_config"] = {"format": {"type": "json_schema", "schema": schema}}
    try:
        resp = _client.messages.create(**kwargs)
    except anthropic.AuthenticationError:
        raise AuthError()
    except anthropic.RateLimitError:
        raise RateLimitHit("APIレート制限")
    except anthropic.APIConnectionError:
        raise ConnError()
    except TypeError as e:
        if "output_config" in str(e):
            sys.exit("[エラー] anthropic パッケージが古く構造化出力に未対応です。"
                     "pip install -U anthropic で更新してください")
        raise
    if resp.stop_reason == "max_tokens":
        raise RuntimeError("応答がトークン上限で途切れました（settings.MAX_TOKENS を増やしてください）")
    if resp.stop_reason == "refusal":
        raise RuntimeError("AIが判定を辞退しました（refusal）")
    text = next((b.text for b in resp.content if b.type == "text"), "")
    return text, resp.usage


_claude_exe = None  # main で resolve_claude により解決


def resolve_claude(cmd, verbose=False):
    """claude 実行ファイルを探す。PATH に無くても、npm のグローバル導入先など
    Windows の既定の場所まで見に行く（VS Code 等 PATH が違う環境でも動くように）。
    見つからなければ None。verbose=True のとき、失敗時に探索先を表示する。
    設定に %APPDATA% などの環境変数を書いてもよい（展開してから探す）。"""
    cmd = os.path.expandvars(cmd or "claude")
    tried = []

    def hit(path):
        ok = bool(path) and os.path.exists(path)
        tried.append((path, ok))
        return ok

    # 1) PATH から（拡張子付きも試す）
    for name in (cmd, cmd + ".cmd", cmd + ".exe", cmd + ".bat"):
        p = shutil.which(name)
        if p:
            return p
    # 2) フルパス指定（設定に絶対パス／環境変数を書いた場合）
    if os.path.isabs(cmd):
        base_noext = os.path.splitext(cmd)[0]
        for c in (cmd, base_noext + ".cmd", base_noext + ".exe", base_noext + ".bat"):
            if hit(c):
                return c
        name = os.path.basename(base_noext)   # 以降は基底名で既定の場所を探す
    else:
        name = cmd
    # 3) npm / node の既定location を直接確認
    for base in (os.environ.get("APPDATA"), os.environ.get("LOCALAPPDATA"),
                 os.environ.get("ProgramFiles"), os.environ.get("ProgramFiles(x86)")):
        if not base:
            continue
        for sub in ("npm", "nodejs", os.path.join("npm", "node_modules", ".bin"), ""):
            for ext in (".cmd", ".exe", ".bat", ""):
                cand = os.path.join(base, sub, name + ext) if sub else os.path.join(base, name + ext)
                if hit(cand):
                    return cand
    # 4) 実シェルに問い合わせる最終手段（where / npm config get prefix）
    for probe in (["where", name], ["npm", "config", "get", "prefix"]):
        try:
            out = subprocess.run(["cmd", "/c"] + probe, capture_output=True,
                                 text=True, encoding="utf-8", errors="replace", timeout=10).stdout
        except Exception:
            out = ""
        for line in (out or "").splitlines():
            line = line.strip()
            if not line:
                continue
            if probe[0] == "where":
                if hit(line):
                    return line
            else:  # npm prefix ディレクトリ配下を確認
                for ext in (".cmd", ".exe", ".bat", ""):
                    if hit(os.path.join(line, name + ext)):
                        return os.path.join(line, name + ext)
    if verbose and tried:
        print("  [参考] 次の場所を確認しましたが見つかりませんでした:")
        for path, ok in tried:
            print(f"     {'あり' if ok else 'なし'} : {path}")
    return None


def _cli_argv(model):
    """claude -p の起動引数を組む。Windowsの .cmd は cmd /c 経由で確実に起動する"""
    exe = _claude_exe or CLAUDE_CMD
    argv = [exe, "-p", "--model", _cli_model(model), "--output-format", "json"]
    if os.name == "nt" and str(exe).lower().endswith((".cmd", ".bat")):
        argv = ["cmd", "/c"] + argv
    return argv


def _classify_cli_error(msg):
    low = (msg or "").lower()
    if any(k in low for k in ("rate limit", "usage limit", "limit reached", "429",
                              "使用上限", "上限に達", "上限を超")):
        return RateLimitHit(msg[:200])
    if any(k in low for k in ("login", "unauthor", "authenticat", "ログイン", "not logged")):
        return AuthError()
    return None


def _cli_call(model, system_text, user_text, want_json):
    """Claude Code をヘッドレス（claude -p）で1回呼ぶ。サブスク認証で動く。
    プロンプトは長くなるため標準入力(stdin)で渡す（引数長の上限を避ける）。"""
    prompt = system_text + "\n\n" + user_text
    if want_json:
        prompt += ("\n\n重要: 回答は次のキーを持つJSONオブジェクト1個だけを出力してください"
                   "（前後に説明文やコードフェンスを付けない）: "
                   "action(要/静観/不要), impact(大/中/小), confidence(高/中/低), reason, flow_note。")
    try:
        proc = subprocess.run(
            _cli_argv(model), input=prompt, capture_output=True, text=True,
            encoding="utf-8", errors="replace", timeout=CLI_TIMEOUT)
    except FileNotFoundError:
        raise ConnError()
    except subprocess.TimeoutExpired:
        raise RuntimeError(f"claude 実行がタイムアウト（{CLI_TIMEOUT}秒）。settings.CLI_TIMEOUT を増やしてください")

    if proc.returncode != 0:
        err = (proc.stderr or proc.stdout or "").strip()
        raise _classify_cli_error(err) or RuntimeError("claude 実行に失敗: " + (err[:200] or "不明なエラー"))

    # 出力例: {"type":"result","subtype":"success","is_error":false,
    #          "result":"…本文…","usage":{...},"total_cost_usd":0.02}
    try:
        wrapper = json.loads(proc.stdout)
    except json.JSONDecodeError:
        return proc.stdout, _Usage()
    text = wrapper.get("result", "")
    u = wrapper.get("usage") or {}
    usage = _Usage(u.get("input_tokens", 0), u.get("output_tokens", 0),
                   u.get("cache_read_input_tokens", 0), u.get("cache_creation_input_tokens", 0),
                   wrapper.get("total_cost_usd"))
    if wrapper.get("is_error") or wrapper.get("subtype") not in (None, "success"):
        raise _classify_cli_error(text or wrapper.get("subtype", "")) or \
            RuntimeError("claude 応答がエラー: " + (str(text)[:200] or wrapper.get("subtype", "")))
    return text, usage


def llm_json(model, system_text, user_text):
    """Claude を呼び、判定JSON（辞書）を返す。バックエンドは設定で切替"""
    if BACKEND == "claude_code":
        text, usage = _cli_call(model, system_text, user_text, True)
    else:
        text, usage = _api_call(model, system_text, user_text, JUDGE_SCHEMA)
    return _parse_json(text), usage


def llm_text(model, system_text, user_text):
    """Claude を呼び、自由文（要約など）をテキストで返す"""
    if BACKEND == "claude_code":
        text, usage = _cli_call(model, system_text, user_text, False)
    else:
        text, usage = _api_call(model, system_text, user_text, None)
    return (text or "").strip(), usage


def build_user_text(row_info, ctx, similar):
    log = ctx["log_text"]
    if len(log) > MAX_LOG_CHARS:
        log = log[:MAX_LOG_CHARS] + f"\n…（以下省略。全{len(ctx['log_text'])}文字）"
    return (
        "以下のアラートを判定してください。\n\n"
        "## 対象アラート\n"
        f"- 日付(UTC): {row_info['date']}\n"
        f"- 発生時刻(JST): {row_info['time']}\n"
        f"- アラーム名: {ctx['alarm_name']}\n"
        f"- サービス: {ctx['service'] or '不明'} / 環境: {ctx['env'] or '不明'}"
        f" / 種別: {ctx['source_type']}\n\n"
        f"## 取得ログ\n```\n{log}\n```\n\n"
        f"## 過去の類似事例（Backlogより）\n{format_cases(similar)}\n"
    )


# 多観点判定で使う観点（ペルソナ）。確信度「低」の件だけこの3観点で再判定する
PERSONAS = [
    ("業務影響", "業務停止・データ欠損・利用者影響が実際に発生するか（波及するか）を最優先の軸として判定してください。"),
    ("復旧・再処理", "自動リトライ・再処理・次サイクルでの自然回復の仕組みで解消するか（自動救済の有無）を最優先の軸として判定してください。"),
    ("事例照合", "過去の類似事例で運用チームがどう判断してきたかとの整合を最優先の軸として判定してください。"),
]

_SEVERITY = {"大": 3, "中": 2, "小": 1}


def persona_suffix(name, instruction):
    """観点指定を user メッセージ末尾に付ける（システム側のキャッシュを壊さないため）"""
    return (
        f"\n\n## 判定観点の指定\n"
        f"あなたは今回「{name}」の観点に立つ判定者です。{instruction}\n"
        f"この観点を最優先しつつ、総合的に判定してください。"
    )


def aggregate_personas(p_results):
    """観点別の判定を多数決で集約する。2観点以上一致すれば採用、割れたら None"""
    if len(p_results) < 2:
        return None
    votes = {}
    for _, r in p_results:
        votes[r["action"]] = votes.get(r["action"], 0) + 1
    action, n = max(votes.items(), key=lambda kv: kv[1])
    if n < 2:
        return None  # 3観点バラバラ → 上位モデルに委ねる
    agree = [(name, r) for name, r in p_results if r["action"] == action]
    # 影響度は一致した観点の中で最も重いものを採用（安全側）
    impact = max((r["impact"] for _, r in agree), key=lambda x: _SEVERITY.get(x, 0))
    reasons = " / ".join(f"【{name}】{r['reason']}" for name, r in agree)
    flow = next((r["flow_note"] for _, r in agree
                 if r.get("flow_note") and r["flow_note"] != "特になし"), "特になし")
    return {
        "action": action, "impact": impact,
        "confidence": "高" if n == len(p_results) else "中",
        "reason": f"{n}/{len(p_results)}観点一致: {reasons}"[:1500],
        "flow_note": flow,
    }


def _trunc_log(ctx):
    log = ctx["log_text"]
    if len(log) > MAX_LOG_CHARS:
        return log[:MAX_LOG_CHARS] + f"\n…（以下省略。全{len(ctx['log_text'])}文字）"
    return log


def judge_pipeline(ctx, row_info, similar, sys_text, add_usage, bump):
    """案A: Sonnet判定 →（確信度低なら）多観点多数決 → Opus再判定"""
    user_text = build_user_text(row_info, ctx, similar)
    result, u = llm_json(MODEL_JUDGE, sys_text, user_text)
    add_usage(u)
    method = f"AI:{MODEL_JUDGE}"
    bump("AI")

    # 確信度が低ければ複数観点で再判定して多数決
    p_results = []
    if result["confidence"] == "低" and MULTI_PERSONA:
        for pname, pinst in PERSONAS:
            try:
                pr, pu = llm_json(MODEL_JUDGE, sys_text, user_text + persona_suffix(pname, pinst))
                add_usage(pu)
                p_results.append((pname, pr))
            except (RateLimitHit, AuthError, ConnError):
                raise  # 致命的なものは行の失敗として上げる
            except Exception:
                pass  # 一部観点の失敗は残りで多数決
        agg = aggregate_personas(p_results)
        if agg:
            result = agg
            method = f"AI多観点:{MODEL_JUDGE}×{len(p_results)}"
            bump("多観点")

    # それでも低い（多数決不成立含む）なら上位モデルで最終判定
    if result["confidence"] == "低" and ESCALATE_ON_LOW_CONFIDENCE:
        esc = user_text
        if p_results:
            split = "\n".join(f"- 【{n}】{r['action']}/{r['impact']}: {r['reason'][:150]}"
                              for n, r in p_results)
            esc += (f"\n\n## 参考: 観点別の判定（多数決不成立）\n{split}\n"
                    "観点間で判断が割れています。割れた理由を踏まえて最終判定してください。")
        try:
            result, u2 = llm_json(MODEL_ESCALATION, sys_text, esc)
            add_usage(u2)
            method = f"AI再判定:{MODEL_ESCALATION}"
            bump("エスカレーション")
        except (RateLimitHit, AuthError, ConnError):
            raise
        except Exception:
            pass  # 再判定に失敗しても一次判定を採用
    return result, method


def judge_orchestration(ctx, row_info, similar, guidelines, system_context, add_usage, bump):
    """案B: 下位モデル（Haiku/Sonnet）が照合・読取・要約 → Opusがその要約で判断"""
    log = _trunc_log(ctx)
    alarm_line = (f"アラーム名: {ctx['alarm_name']} / サービス: {ctx['service'] or '不明'}"
                  f" / 環境: {ctx['env'] or '不明'} / 種別: {ctx['source_type']}")

    # 作業者1: 過去事例の要約
    w1, u = llm_text(
        MODEL_WORKER,
        "あなたは運用アラート判定の補助担当です。過去事例を、判定の材料として簡潔に要約します。",
        f"## 対象ログ\n{log}\n\n## 過去の類似事例\n{format_cases(similar)}\n\n"
        "この類似事例から、今回のログに対して過去どう判断されてきたか（対応要否の傾向と主な理由）を3〜5行で要約してください。")
    add_usage(u)

    # 作業者2: 判定基準の該当部抽出
    w2, u = llm_text(
        MODEL_WORKER,
        "あなたは運用アラート判定の補助担当です。次の判定基準から該当箇所を抜き出します。\n\n# 判定基準\n" + guidelines,
        f"## 対象ログ\n{log}\n\nこのアラーム／ログに関係する判定基準の該当箇所"
        "（静観してよい条件・対応が要る条件・影響度の目安）を、要点だけ箇条書きで抽出してください。")
    add_usage(u)

    # 作業者3: システム構成の関連フロー
    w3, u = llm_text(
        MODEL_WORKER,
        "あなたは運用アラート判定の補助担当です。次のシステム背景を参照します。\n\n# システム背景\n" + system_context,
        f"## 対象アラート\n{alarm_line}\n## 対象ログ\n{log}\n\n"
        "このアラームについて、(1)上流起因のサインがあるか (2)下流へ波及するか "
        "(3)次サイクルで自然回復するか を、構成に照らして3〜5行で要約してください。")
    add_usage(u)

    # 指揮官（Opus）: 要約だけを読んで最終判定（大きなファイル本文は読まない＝費用を抑える）
    judge_user = (
        "以下は下位モデルが判定基準・システム背景・過去事例を読み取って要約した材料です。"
        "これらと対象ログをもとに最終判定してください。\n\n"
        f"## 対象アラート\n{alarm_line}\n\n## 対象ログ\n```\n{log}\n```\n\n"
        f"## 過去事例の傾向（要約）\n{w1}\n\n"
        f"## 関係する判定基準（要約）\n{w2}\n\n"
        f"## システム構成の所見（要約）\n{w3}\n")
    result, u = llm_json(MODEL_ORCHESTRATOR, ROLE_TEXT, judge_user)
    add_usage(u)
    bump("AI")
    bump("オーケストレーション")
    method = f"AI指揮:{MODEL_ORCHESTRATOR}/作業:{MODEL_WORKER}"
    return result, method


# ============================================================
# シート・行の処理
# ============================================================

def pick_sheet(wb):
    if SHEET_NAME:
        if SHEET_NAME not in wb.sheetnames:
            sys.exit(f"[エラー] シート「{SHEET_NAME}」がありません。存在するシート: {wb.sheetnames}")
        return wb[SHEET_NAME]
    this_month = datetime.date.today().strftime("%Y-%m")
    if this_month in wb.sheetnames:
        return wb[this_month]
    print(f"[注意] 実行月のシート({this_month})が無いため先頭シート「{wb.sheetnames[0]}」を使います")
    return wb[wb.sheetnames[0]]


def target_rows(ws):
    """判定対象の行番号を返す。
    対象 = D列にログがあり、「【取得できず】」ではなく、（再判定でなければ）G列が空の行"""
    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        errors = ws.cell(r, COL_ERRORS).value
        if errors is None or str(errors).strip() == "":
            continue
        if str(errors).startswith("【取得できず】"):
            continue
        if not RE_JUDGE and ws.cell(r, COL_AI_ACTION).value not in (None, ""):
            continue
        rows.append(r)
    return rows


# 理由の段落ラベル（この語の直前で必ず改行し、ブロック間に空行を入れる）
REASON_LABELS = ("結論", "根拠", "確認事項", "前後関係")


def format_reason(text):
    """理由を読みやすく整形する。
      - 【結論】【根拠】【確認事項】【前後関係】の各ブロックの前で改行し、
        ブロック間には空行を1行入れる
      - 各ブロック内も文（句点「。」区切り）ごとに改行して1文1行にする"""
    s = str(text)
    # 各ブロックラベルの直前で必ず改行（ブロック境界を確実に分ける）
    pat = r"\s*(?=【(?:" + "|".join(REASON_LABELS) + r")】)"
    s = re.sub(pat, "\n", s).lstrip("\n")

    out = []
    for line in s.split("\n"):
        if line.strip() == "":
            continue
        # 「。」の直後で分割（末尾の空要素は捨てる）。文が無ければ元の行のまま
        parts = [p.strip() for p in re.split(r"(?<=。)", line) if p.strip() != ""]
        seg = parts if parts else [line]
        # ブロックの先頭（【…】で始まる）の前に空行を入れて区切る
        if out and seg and seg[0].startswith("【"):
            out.append("")
        out.extend(seg)
    return "\n".join(out)


# 差分記録ファイル（永続）の列。運用チームの手入力列（運用対応要否・差分理由メモ）は
# ツールが上書きしない。差分列は Excel 数式で自動判定する。
DIFF_COLS = ["日付", "アラーム名", "時刻",
             "AI対応要否", "AI影響度", "AI確信度", "AI判定日時",
             "運用対応要否", "差分", "差分理由・改善メモ"]


def _diff_key(date, alarm, time_):
    return "|".join(str(x if x is not None else "").strip() for x in (date, alarm, time_))


def update_diff_record(path, sheet_name, records):
    """AI判定を差分記録ファイルへ upsert（追記/更新）する。
    キー（日付＋アラーム名＋時刻）が既にあれば AI列だけ更新し、
    運用チームが記入した列（運用対応要否・差分理由メモ）は保持する。
    差分列は数式で自動判定（運用対応要否が入力されると「一致/差分」を表示）。
    返り値: (新規件数, 更新件数)"""
    if not records:
        return (0, 0)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # 見出しが未作成なら作る
    if ws.cell(1, 1).value in (None, ""):
        for c, name in enumerate(DIFF_COLS, 1):
            ws.cell(1, c).value = name
        ws.column_dimensions[get_column_letter(2)].width = 42   # アラーム名
        ws.column_dimensions[get_column_letter(9)].width = 8    # 差分
        ws.column_dimensions[get_column_letter(10)].width = 50  # 差分理由メモ

    # 既存キー → 行番号
    index = {}
    for r in range(2, ws.max_row + 1):
        k = _diff_key(ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
        if k.strip("|"):
            index[k] = r

    n_new = n_upd = 0
    for rec in records:
        k = _diff_key(rec["date"], rec["alarm"], rec["time"])
        r = index.get(k)
        if r is None:
            r = ws.max_row + 1
            index[k] = r
            ws.cell(r, 1).value = rec["date"]
            ws.cell(r, 2).value = rec["alarm"]
            ws.cell(r, 3).value = rec["time"]
            n_new += 1
        else:
            n_upd += 1
        # AI列は毎回更新
        ws.cell(r, 4).value = rec["action"]
        ws.cell(r, 5).value = rec["impact"]
        ws.cell(r, 6).value = rec["confidence"]
        ws.cell(r, 7).value = rec["judged_at"]
        # 差分列（数式）。運用対応要否(H)が入ると一致/差分を自動表示
        ws.cell(r, 9).value = f'=IF($H{r}="","",IF($D{r}=$H{r},"一致","差分"))'
        # 運用対応要否(8)・差分理由メモ(10) は手入力なので触らない
    wb.save(path)
    return (n_new, n_upd)


def write_result(ws, r, action, impact, reason, confidence, method):
    ws.cell(r, COL_AI_ACTION).value = action
    ws.cell(r, COL_AI_IMPACT).value = impact
    cell = ws.cell(r, COL_AI_REASON)
    cell.value = reason
    # 理由は長文なのでセル内で折り返し・上揃えにして読みやすくする
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(r, COL_AI_CONF).value = confidence
    ws.cell(r, COL_AI_METHOD).value = method


def write_headers(ws):
    """G〜K列に見出しを書く。想定外の見出しが既にある場合は
    シートの別用途の列を上書きしないよう中止する"""
    headers = {COL_AI_ACTION: "AI対応要否", COL_AI_IMPACT: "AI影響度",
               COL_AI_REASON: "AI判定理由", COL_AI_CONF: "AI確信度",
               COL_AI_METHOD: "判定方式"}
    for col, name in headers.items():
        cur = ws.cell(HEADER_ROW, col).value
        if cur in (None, ""):
            ws.cell(HEADER_ROW, col).value = name
        elif str(cur).strip() != name:
            sys.exit(
                f"[エラー] {get_column_letter(col)}1 に想定外の見出し「{cur}」があります。\n"
                f"  G〜K列は本ツールの書き込み先です。シート側で別の用途に使われている場合、\n"
                f"  上書きを防ぐため中止します（列の割り当てを見直してください）"
            )
    # 理由(I列)は長文なので幅を広めに確保。他の判定列も見出しに合わせて整える
    ws.column_dimensions[get_column_letter(COL_AI_REASON)].width = 70
    for col in (COL_AI_ACTION, COL_AI_IMPACT, COL_AI_CONF):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(COL_AI_METHOD)].width = 32


# ============================================================
# メイン
# ============================================================

def main():
    print("=== アラートログ判定ツール ===")

    # --- 入力の確認 ---
    if not os.path.exists(INPUT_XLSX):
        sys.exit(f"[エラー] 入力ファイルが見つかりません: {INPUT_XLSX}")

    # --- 出力パスの決定（入力と同じパスなら中止） ---
    out_dir = OUTPUT_DIR or os.path.dirname(INPUT_XLSX)
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(INPUT_XLSX))[0]
    out_path = os.path.join(out_dir, OUTPUT_XLSX or f"{base}_判定済.xlsx")
    if os.path.abspath(out_path) == os.path.abspath(INPUT_XLSX):
        sys.exit("[エラー] 出力先が入力ファイルと同じです。上書きを防ぐため中止します")

    # --- ナレッジ読み込み ---
    rules, cases, guidelines, system_context = load_knowledge()
    print(f"ナレッジ読込: ルール{len(rules.get('rules', []))}件 / 過去事例{len(cases)}件")

    # --- バックエンド準備（DRY_RUN なら不要） ---
    global _client, _claude_exe
    if not DRY_RUN:
        if BACKEND == "claude_code":
            _claude_exe = resolve_claude(CLAUDE_CMD, verbose=True)
            if not _claude_exe:
                sys.exit(
                    f"[エラー] Claude Code（{CLAUDE_CMD}）が見つかりませんでした。\n"
                    "  PATH に無い環境（VS Code など）ではこのエラーになります。次のいずれかで解決します。\n"
                    "  1) settings.py の CLAUDE_CMD に claude.cmd の場所を指定する（環境変数も使えます）\n"
                    "       例）CLAUDE_CMD = r\"%APPDATA%\\npm\\claude.cmd\"\n"
                    "  2) 未導入なら:  npm install -g @anthropic-ai/claude-code  → 一度 `claude` でログイン\n"
                    "  ※AIを呼ばずに動作だけ試すなら settings.py の DRY_RUN = True"
                )
            print(f"Claude Code: {_claude_exe}")
        else:  # api
            if not os.environ.get("ANTHROPIC_API_KEY"):
                sys.exit(
                    "[エラー] 環境変数 ANTHROPIC_API_KEY が設定されていません。\n"
                    '  PowerShell:  $env:ANTHROPIC_API_KEY = "sk-ant-..."\n'
                    "  ※サブスク運用なら settings.py の BACKEND = \"claude_code\"\n"
                    "  ※お試しだけなら settings.py の DRY_RUN = True で動きます"
                )
            import anthropic
            _client = anthropic.Anthropic()

    # 案A用のシステムプロンプト（役割＋判定基準＋システム背景。毎回同じでキャッシュが効く）
    pipeline_sys = build_pipeline_system(guidelines, system_context)

    # --- ブック読み込み → コピーに書き込み ---
    try:
        shutil.copyfile(INPUT_XLSX, out_path)
    except PermissionError:
        sys.exit(f"[エラー] 出力先に書き込めません。前回の出力ファイルをExcelで開いていたら閉じてください: {out_path}")
    wb = openpyxl.load_workbook(out_path)
    ws = pick_sheet(wb)
    write_headers(ws)

    rows = target_rows(ws)
    if len(rows) > MAX_ROWS:
        print(f"[注意] 対象{len(rows)}行のうち先頭{MAX_ROWS}行だけ処理します（settings.MAX_ROWS）")
        rows = rows[:MAX_ROWS]
    if DRY_RUN:
        mode = "DRY_RUN(機械判定のみ)"
    else:
        be = "サブスク(claude -p)" if BACKEND == "claude_code" else "API"
        jm = "オーケストレーション(案B)" if JUDGE_MODE == "orchestration" else "パイプライン(案A)"
        mode = f"{be} / {jm}"
    print(f"シート「{ws.title}」 判定対象: {len(rows)}行 / モード: {mode}")

    # --- 集計用 ---
    stats = {"機械": 0, "AI": 0, "多観点": 0, "オーケストレーション": 0,
             "エスカレーション": 0, "AI対象(未実施)": 0, "失敗": 0}
    usage_total = {"in": 0, "out": 0, "cache_read": 0, "cache_write": 0, "cost": 0.0}

    def add_usage(u):
        usage_total["in"] += getattr(u, "input_tokens", 0) or 0
        usage_total["out"] += getattr(u, "output_tokens", 0) or 0
        usage_total["cache_read"] += getattr(u, "cache_read_input_tokens", 0) or 0
        usage_total["cache_write"] += getattr(u, "cache_creation_input_tokens", 0) or 0
        if getattr(u, "cost", None):
            usage_total["cost"] += u.cost

    def bump(k):
        stats[k] = stats.get(k, 0) + 1

    # 差分記録用（AI判定を別ファイルへ蓄積する材料）
    diff_records = []
    judged_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    def record(r, action, impact, confidence):
        diff_records.append({
            "date": ws.cell(r, COL_DATE).value,
            "alarm": str(ws.cell(r, COL_ALARM).value or ""),
            "time": ws.cell(r, COL_TIME).value,
            "action": action, "impact": impact, "confidence": confidence,
            "judged_at": judged_at,
        })

    for i, r in enumerate(rows, 1):
        alarm = str(ws.cell(r, COL_ALARM).value or "")
        log_text = str(ws.cell(r, COL_ERRORS).value or "")
        row_info = {"date": ws.cell(r, COL_DATE).value, "time": ws.cell(r, COL_TIME).value}
        ctx = alarm_context(alarm, log_text)
        label = f"[{i}/{len(rows)}] 行{r} {alarm[:50]}"

        # 1) 機械判定（rules.yaml）
        rule = apply_rules(rules, ctx)
        if rule:
            then = rule.get("then", {})
            note = f"（{rule.get('note')}）" if rule.get("note") else ""
            write_result(ws, r,
                         then.get("action", "要確認"), then.get("impact", ""),
                         f"{then.get('reason','')}{note}", "確定",
                         f"機械:{rule.get('name','')}")
            record(r, then.get("action", "要確認"), then.get("impact", ""), "確定")
            stats["機械"] += 1
            print(f"{label} → 機械判定: {then.get('action','')}")
            continue

        # 2) 類似事例検索
        similar = find_similar_cases(ctx, cases, TOP_N_CASES)

        if DRY_RUN:
            stats["AI対象(未実施)"] += 1
            print(f"{label} → AI判定対象（類似事例{len(similar)}件）")
            continue

        # 3) AI判定（方式で分岐）
        try:
            if JUDGE_MODE == "orchestration":
                result, method = judge_orchestration(
                    ctx, row_info, similar, guidelines, system_context, add_usage, bump)
            else:
                result, method = judge_pipeline(
                    ctx, row_info, similar, pipeline_sys, add_usage, bump)

            reason = result["reason"]
            if result.get("flow_note") and result["flow_note"] != "特になし":
                reason += f"\n【前後関係】{result['flow_note']}"
            reason = format_reason(reason)
            write_result(ws, r, result["action"], result["impact"],
                         reason, result["confidence"], method)
            record(r, result["action"], result["impact"], result["confidence"])
            print(f"{label} → {result['action']}/{result['impact']} (確信度:{result['confidence']})")

        except AuthError:
            sys.exit("[エラー] 認証に失敗しました。\n"
                     "  API: ANTHROPIC_API_KEY を確認 / サブスク: `claude` でログイン済みか確認してください")
        except RateLimitHit as e:
            write_result(ws, r, "", "", f"使用上限のため未判定（時間をおいて再実行）: {e}", "", "失敗")
            bump("失敗")
            print(f"{label} → 失敗（使用上限）")
        except ConnError:
            sys.exit("[エラー] 接続できません。\n"
                     "  API: 社内プロキシ配下なら pip install truststore を確認\n"
                     "  サブスク: Claude Code の導入・ログインを確認してください")
        except Exception as e:
            write_result(ws, r, "", "", f"判定失敗: {e}", "", "失敗")
            bump("失敗")
            print(f"{label} → 失敗: {e}")

    # --- 保存（Excelで開きっぱなしの場合に親切に案内） ---
    try:
        wb.save(out_path)
    except PermissionError:
        sys.exit(f"[エラー] 保存できません。出力ファイルをExcelで開いていたら閉じてください: {out_path}")

    # --- 差分記録ファイルへ反映（別ファイル・永続。運用チームの記入列は保持） ---
    diff_msg = ""
    if DIFF_RECORD and not DRY_RUN and diff_records:
        try:
            n_new, n_upd = update_diff_record(DIFF_RECORD_XLSX, ws.title, diff_records)
            diff_msg = f"  差分記録: 新規{n_new}件 / 更新{n_upd}件 → {DIFF_RECORD_XLSX}"
        except PermissionError:
            diff_msg = f"  [注意] 差分記録ファイルを開いていたため更新できませんでした（閉じて再実行）: {DIFF_RECORD_XLSX}"

    # --- 結果サマリ ---
    print("\n=== 結果 ===")
    if not DRY_RUN:
        be = "Claude Code(サブスク)" if BACKEND == "claude_code" else "API"
        jm = "オーケストレーション(案B)" if JUDGE_MODE == "orchestration" else "パイプライン(案A)"
        print(f"  バックエンド: {be} / 方式: {jm}")
    for k, v in stats.items():
        if v:
            print(f"  {k}: {v}件")
    if not DRY_RUN:
        if usage_total["in"] or usage_total["out"]:
            print(f"  トークン: 入力{usage_total['in']:,} / 出力{usage_total['out']:,}"
                  f" / キャッシュ読み{usage_total['cache_read']:,} / キャッシュ書き{usage_total['cache_write']:,}")
            if BACKEND == "api" and usage_total["cache_read"] == 0 and stats.get("AI", 0) >= 2:
                print("  [注意] キャッシュが効いていません（想定外。設定変更があれば見直してください）")
        if usage_total["cost"]:
            print(f"  概算コスト: ${usage_total['cost']:.4f}")
        if BACKEND == "claude_code":
            print("  ※ サブスク(claude -p)経由のため費用はプランに含まれ、使用枠を消費します"
                  "（案Bは1件で複数回呼ぶため枠消費が増えます）")
    print(f"  出力: {out_path}")
    if diff_msg:
        print(diff_msg)
    print("※ AI判定は参考情報です。最終判断（E列）は運用チームの目視で行ってください")


if __name__ == "__main__":
    main()
