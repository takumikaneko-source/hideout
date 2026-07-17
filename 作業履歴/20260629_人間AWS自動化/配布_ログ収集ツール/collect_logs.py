# -*- coding: utf-8 -*-
"""
ログ収集ツール本体（運用チーム向け）。

【このツールがすること】
  1. ダウンロードした Slack通知アラート収集スプレッドシート(.xlsx)を読む。
  2. まだログを取っていない行（D列が空の行）を対象に、
     A列の日時・B列のアラーム名から、対応する CloudWatch のログを取得する。
  3. 入力ファイルの「コピー」を作り、そのコピーの
     D列（ログ本文）と F列（AWSコンソールへのリンク）に結果を書き込む。
  4. 運用担当者は、その D/F セルを本物のスプレッドシートの該当行に貼り付ける。

【処理後の D列の意味（結果の見方）】
  ・ログ本文が入っている … 取得できた（貼り付け対象）。
  ・「【該当ログなし】」   … その時間帯にログが無かった。
  ・「【取得できず】…」   … 未対応や設定不備など。理由が書いてあるので確認する。
  ・空のまま             … 一時的なエラー。もう一度実行すれば取り直す。

【設計方針】
  ・設定は settings.py に分けてある。運用では settings.py だけを編集し、
    このファイル(collect_logs.py)は変更しない。
  ・AWSの資格情報(アクセスキー等)は、このスクリプトからは読み書きしない
    （boto3 が実行時に各自の ~/.aws を参照する）。
  ・AWSに対しては「読み取り」しか行わない（作成・変更・削除はしない）。
  ・ログの判定にAIは使わない（このツールは取得と転記だけ）。
"""

import os
import sys
import shutil
from datetime import datetime, timezone, timedelta
from urllib.parse import quote

# ---- 設定の既定値 -----------------------------------------------------------
# ここは触らない。実際の設定は settings.py で行う。
# settings.py が無い、または一部の項目が書かれていない場合に、この既定値が使われる。
INPUT_XLSX = ""
SHEET_NAME = ""
OUTPUT_DIR = ""
OUTPUT_XLSX = ""
REGION = "ap-northeast-1"
LINES_BEFORE = 40          # 発火行の前の文脈行数（控えめに。全文は F列リンクで見られる）
LINES_AFTER = 40           # 発火行の後の文脈行数
ERROR_HEADER_MAX = 8       # D列の先頭に再掲する「発火した行」の最大件数
LINE_MAX_CHARS = 800       # 1行が長すぎる（巨大JSON等）ときの1行あたり上限文字数
WINDOW_MIN = 5
MAX_ROWS = 0
PREFIX_TO_PROFILE = {}
DEFAULT_PROFILE = ""
CA_BUNDLE = ""

# ---- settings.py を読み込んで、上の既定値を上書きする -----------------------
# このファイルと同じフォルダにある settings.py を読み込む。
#   sys.path.insert(...) は「settings.py をこのファイルの隣から確実に探す」ための指定。
#   （こうしておくと、別のフォルダからツールを実行しても settings.py を見つけられる）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    import settings as _s
    # settings.py に書かれている項目だけを、上の既定値に上書きする
    for _k in ("INPUT_XLSX", "SHEET_NAME", "OUTPUT_DIR", "OUTPUT_XLSX", "REGION",
               "LINES_BEFORE", "LINES_AFTER", "ERROR_HEADER_MAX", "LINE_MAX_CHARS",
               "WINDOW_MIN", "MAX_ROWS",
               "PREFIX_TO_PROFILE", "DEFAULT_PROFILE", "CA_BUNDLE"):
        if hasattr(_s, _k):
            globals()[_k] = getattr(_s, _k)
except ImportError:
    # settings.py が見つからなくても、既定値で動けるようにしておく（安全側）
    print("[WARN] settings.py が見つかりません。既定値で動作します。", flush=True)

# Excelの1セルに入れられる文字数の上限(32767)より少し手前で切る。
# これより長いログは途中で切り、全文は F列のリンクから見てもらう。
XLSX_CELL_LIMIT = 32000

# スプレッドシートの列番号（1始まり）。A列=1, B列=2, ... の対応。
#   A=日時 / B=アラーム名 / D=ログ本文の書き込み先 / F=リンクの書き込み先
COL_A, COL_B, COL_D, COL_F = 1, 2, 4, 6


def log(msg=""):
    """画面(コンソール)に1行表示する。flush=True で即座に出す（進捗が見えるように）。"""
    print(msg, flush=True)


def setup_ssl():
    """社内ネットワークのSSL検査(プロキシ)による証明書エラーを避けるための準備。

    社内ネットワークは通信を検査するため独自の証明書を挟むことがあり、
    そのままだと AWS への通信が「証明書エラー」で失敗することがある。その対策。
    """
    if CA_BUNDLE:
        # 社内CA(.pem)のパスが settings.py で指定されていれば、それを使う
        os.environ["AWS_CA_BUNDLE"] = CA_BUNDLE
        os.environ["REQUESTS_CA_BUNDLE"] = CA_BUNDLE
        log(f"[SSL ] CA_BUNDLE を使用: {CA_BUNDLE}")
    else:
        # 指定が無ければ、truststore を使って「OSに入っている証明書」で検証する
        try:
            import truststore
            truststore.inject_into_ssl()
            log("[SSL ] truststore: OSの証明書ストアを使用")
        except ImportError:
            # truststore が未インストールでも処理は続ける（証明書エラー時のみ対処が必要）
            log("[SSL ] truststore無し。SSLで失敗する場合は `pip install truststore` か CA_BUNDLE指定")


def parse_utc(val):
    """A列の日時を UTC の日時オブジェクトに変換する。

    セルの値が日時型ならそのまま、文字列なら決まった形式で読み取る。
    どの形式にも当てはまらない（＝解釈できない）場合は None を返す。
    """
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val.astimezone(timezone.utc)
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def ms(dt):
    """日時を「エポックミリ秒」に変換する。CloudWatch のAPIが時刻をミリ秒で受け取るため。"""
    return int(dt.timestamp() * 1000)


def fmt_ts(epoch_ms):
    """エポックミリ秒を「YYYY-MM-DD HH:MM:SS」の読める文字列(UTC)に戻す。表示・ログ用。"""
    return datetime.fromtimestamp(epoch_ms / 1000, timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def console_link(region, log_group, log_stream, start_ms, end_ms):
    """該当ログを AWSコンソールで直接開けるURL(ディープリンク)を組み立てる。

    F列に貼る「リンク」の飛び先。AWSコンソールは記号を独特な形でエンコードするため、
    enc() で二重にエンコードし、'%' を '$' に置き換えている（AWS側の仕様に合わせたもの）。
    """
    def enc(s):
        # AWSコンソールのURL用エンコード（二重エンコードして % を $ に置換）
        return quote(quote(s, safe=""), safe="").replace("%", "$")
    base = f"https://{region}.console.aws.amazon.com/cloudwatch/home?region={region}#logsV2:log-groups"
    url = f"{base}/log-group/{enc(log_group)}/log-events/{enc(log_stream)}"
    # $3F='?', $3D='=', $26='&' に相当（表示する時間範囲を指定）
    url += f"$3FstartTime$3D{start_ms}$26endTime$3D{end_ms}"
    return url


# 再実行しても変わらない（＝要確認）とみなすAWSエラーコード。対象そのものが存在しない類。
_PERMANENT_AWS_CODES = {"ResourceNotFoundException"}


def classify_aws_error(e):
    """AWS例外を (種別, 説明) に分類。
    種別: 'permanent'（再実行しても変わらない＝Dに理由を記録） / 'transient'（再実行で解決しうる＝D空）。
    権限や認証などの全体的なエラーは transient 扱い（設定を直して再実行すれば全行やり直せるため）。"""
    code = ""
    resp = getattr(e, "response", None)
    if isinstance(resp, dict):
        code = resp.get("Error", {}).get("Code", "")
    if code in _PERMANENT_AWS_CODES:
        return "permanent", f"対象が見つかりません（{code}）。ロググループ/ストリームが存在しない可能性があります。"
    return "transient", f"{e!r}"


def resolve_log_group(logs, cw, alarm_name):
    """アラーム名から「どのロググループを見ればよいか」を突き止める。

    戻り値は5つ組: (ロググループ名, 絞り込みパターン, 種類, 補足メモ, 詳細情報)。
      ・ロググループ名が None のときは特定できなかったことを表し、「種類」に理由が入る。
        種類の例: "unknown"=アラーム未検出 / "states"=StepFunctions(未対応) /
                  "unsupported"=未対応の種類 / "mf_error"=一時エラー / "mf_denied"=権限等
    処理の流れ:
      1) アラーム定義を取得（describe_alarms）。
      2) Lambdaのアラームなら、ロググループ名は決まった形 /aws/lambda/<関数名>。
      3) Step Functions は本ツールでは未対応。
      4) それ以外は「メトリクスフィルタ」からロググループを逆引きする。
    """
    da = cw.describe_alarms(AlarmNames=[alarm_name])
    metric_alarms = da.get("MetricAlarms", [])
    if not metric_alarms:
        # 指定した名前のアラームが存在しない（または種類が違う）
        return None, None, "unknown", "アラームが見つからない/メトリクスアラームでない", {}
    a = metric_alarms[0]
    ns = a.get("Namespace")       # 例: "AWS/Lambda" や独自の名前空間
    mn = a.get("MetricName")      # メトリクス名
    dims = {d["Name"]: d["Value"] for d in a.get("Dimensions", [])}  # 付随情報（関数名など）
    detail = {"namespace": ns, "metric": mn, "dimensions": dims}

    # Lambda のアラームは、ロググループ名が「/aws/lambda/<関数名>」で決まっている
    if ns == "AWS/Lambda" and "FunctionName" in dims:
        return f"/aws/lambda/{dims['FunctionName']}", "", "lambda_native", "AWS/Lambda ネイティブ", detail
    # Step Functions は CloudWatch Logs ではなく実行履歴から取得する（下流の main で処理する）
    if ns == "AWS/States":
        return None, None, "states", "Step Functions（失敗実行から取得）", detail

    # 上記以外は「ログのメトリクスフィルタ」からロググループを逆引きする
    try:
        mf = logs.describe_metric_filters(metricName=mn, metricNamespace=ns)
        filters = mf.get("metricFilters", [])
        if filters:
            # メトリクスフィルタが見つかった → その定義から「ロググループ名」と
            # 「絞り込みパターン(例: [ERROR])」を取り出して返す
            f = filters[0]
            detail["metric_filter"] = f.get("filterName")
            return f.get("logGroupName"), f.get("filterPattern", ""), "metric_filter", f"namespace={ns}", detail
    except Exception as e:
        # 逆引きの途中でAWSエラー。権限不足などの恒久的なものと一時的なものを分ける
        kind, msg = classify_aws_error(e)
        stype = "mf_denied" if kind == "permanent" else "mf_error"
        return None, None, stype, msg, detail

    # メトリクスフィルタも無く、対応する取り方が分からなかった
    return None, None, "unsupported", f"未対応 namespace={ns} metric={mn}", detail


def fetch_logs(logs, log_group, filter_pattern, center_utc):
    """指定のロググループから、対象時刻の前後のログを取り出す。

    戻り値は4つ組: (ログ本文, リンク用の情報, 該当件数, 統計情報)。
      ・該当するログが1件も無いときは、ログ本文の代わりに None を返す。
    処理の流れ:
      1) 対象時刻の前後 WINDOW_MIN 分を検索し、一致したログ（発火行）を見つける。
      2) D列本文は「発火した行を先頭に再掲」→「前後の文脈（控えめ）」の順に組み立てる。
         こうすると、長いログでも先頭に必ず発火行が残り、後工程（AI判定・人の目視）が
         『何で鳴ったか』を確実に読める。全文は F列リンクから参照する。
      3) 一致した行の先頭には ">>> " の印を付けて分かりやすくする。
    """
    # 検索する時間の範囲（対象時刻の前後 WINDOW_MIN 分）
    start = center_utc - timedelta(minutes=WINDOW_MIN)
    end = center_utc + timedelta(minutes=WINDOW_MIN)

    # まず範囲内でログを検索する（filter_pattern があれば、その語で絞り込む）
    kwargs = dict(logGroupName=log_group, startTime=ms(start), endTime=ms(end), limit=50)
    if filter_pattern:
        kwargs["filterPattern"] = filter_pattern
    fl = logs.filter_log_events(**kwargs)
    events = fl.get("events", [])
    stat = {"window": f"{fmt_ts(ms(start))}〜{fmt_ts(ms(end))} UTC", "matched": len(events)}
    if not events:
        # 範囲内に一致するログが無かった（＝該当なし）
        return None, None, 0, stat

    # 最初に一致したログを基点にする（そのログが在るストリームと時刻を覚える）
    first = events[0]
    stream = first["logStreamName"]
    hit_ts = first["timestamp"]
    stat["stream"] = stream
    stat["hit_time"] = fmt_ts(hit_ts)

    # 基点のストリームから、前後の連続したログを取得する（前後の文脈を読めるように）
    ge = logs.get_log_events(
        logGroupName=log_group, logStreamName=stream,
        startTime=hit_ts - WINDOW_MIN * 60 * 1000,
        endTime=hit_ts + WINDOW_MIN * 60 * 1000,
        startFromHead=True, limit=10000,
    )
    stream_events = ge.get("events", [])
    stat["stream_events"] = len(stream_events)

    # 取得した並びの中で、基点(hit_ts)の位置を探す
    idx = 0
    for i, e in enumerate(stream_events):
        if e["timestamp"] >= hit_ts:
            idx = i
            break

    # 基点の前 LINES_BEFORE 行〜後 LINES_AFTER 行だけを切り出す（前後の文脈）
    lo = max(0, idx - LINES_BEFORE)
    hi = min(len(stream_events), idx + LINES_AFTER + 1)
    stat["extracted"] = hi - lo

    def _fmt(e, mark="    "):
        """1件を「[時刻] 本文」に整形。1行が長すぎる（巨大JSON等）ときは切る。"""
        msg = e["message"].rstrip()
        if len(msg) > LINE_MAX_CHARS:
            msg = msg[:LINE_MAX_CHARS] + " …(1行が長いため省略)"
        return f"{mark}[{fmt_ts(e['timestamp'])}] {msg}"

    # --- (1) 発火した行を先頭に再掲する ---
    #   検知パターンに一致した行(events)を先頭にまとめて置く。長いログでD列が切られても、
    #   『何で鳴ったか』が必ず残るようにするため（後工程のAI判定・人の目視の両方に効く）。
    header = ["■ 発火した行（アラームの検知パターンに一致）:"]
    for e in events[:ERROR_HEADER_MAX]:
        header.append(_fmt(e, ">>> "))
    if len(events) > ERROR_HEADER_MAX:
        header.append(f"    （ほか {len(events) - ERROR_HEADER_MAX} 件の一致行あり。全文は F列リンク参照）")

    # --- (2) 前後の文脈（控えめに）---
    context = [f"── 前後の文脈（前{LINES_BEFORE}行 / 後{LINES_AFTER}行。全文は F列リンク参照）──"]
    for i in range(lo, hi):
        context.append(_fmt(stream_events[i], ">>> " if i == idx else "    "))

    text = "\n".join(header) + "\n\n" + "\n".join(context)
    link = (log_group, stream, ms(start), ms(end))  # F列リンクを組み立てるための材料
    return text, link, len(events), stat


# ============================================================
# Step Functions 用（CloudWatch Logs ではなく「実行履歴」から失敗内容を取る）
# ============================================================

def sf_console_link(region, execution_arn):
    """Step Functions の実行画面(コンソール)を直接開くリンクを作る。F列に貼る飛び先。"""
    return (f"https://{region}.console.aws.amazon.com/states/home?region={region}"
            f"#/executions/details/{execution_arn}")


def _sf_failure_detail(sfn, execution_arn):
    """1つの実行の履歴(新しい順)をたどって、失敗の error / cause / 失敗ステート名 を取り出す。"""
    error = cause = failed_state = None
    # 失敗イベントの種類ごとに、error/cause が入っている項目名
    fail_detail_keys = ("executionFailedEventDetails", "taskFailedEventDetails",
                        "lambdaFunctionFailedEventDetails", "activityFailedEventDetails",
                        "mapRunFailedEventDetails")
    token = None
    scanned = 0
    while scanned < 1000:   # 念のため上限（暴走防止）
        kwargs = {"executionArn": execution_arn, "reverseOrder": True, "maxResults": 100}
        if token:
            kwargs["nextToken"] = token
        resp = sfn.get_execution_history(**kwargs)
        for ev in resp.get("events", []):
            scanned += 1
            # error / cause を拾う（最初に見つかったものを採用）
            for key in fail_detail_keys:
                d = ev.get(key)
                if d:
                    if error is None:
                        error = d.get("error")
                    if cause is None:
                        cause = d.get("cause")
            # 新しい順にたどって最初に出てくる「ステート開始」が、失敗したステートの手がかり
            if failed_state is None and ev.get("type", "").endswith("StateEntered"):
                sd = ev.get("stateEnteredEventDetails")
                if sd:
                    failed_state = sd.get("name")
            if error is not None and failed_state is not None:
                return error, cause, failed_state
        token = resp.get("nextToken")
        if not token:
            break
    return error, cause, failed_state


def fetch_stepfunctions_error(sfn, region, state_machine_arn, center_utc):
    """Step Functions の「失敗した実行」から失敗内容を取り出す。

    戻り値は3つ組: (本文, 実行画面リンク, 統計情報)。
      ・対象時刻の近くに失敗実行が見つからないときは、本文の代わりに None を返す。
    手順:
      1) 失敗実行を新しい順に取得し、対象時刻の前後（窓）に入るものを集める。
      2) その中から対象時刻に一番近い実行を選ぶ。
      3) その実行の履歴から error / cause / 失敗ステート名 を取り出して整形する。
    """
    # アラームは実行失敗から少し遅れて鳴ることがあるので、窓は広め（最低30分）にする
    window = timedelta(minutes=max(WINDOW_MIN, 30))
    lo = center_utc - window
    hi = center_utc + window
    stat = {"window": f"{fmt_ts(ms(lo))}〜{fmt_ts(ms(hi))} UTC", "scanned": 0}

    # 失敗実行を新しい順に見て、窓に入るものを集める（窓より古くなったら打ち切り）
    candidates = []
    token = None
    while stat["scanned"] < 1000:   # 念のため上限（暴走防止）
        kwargs = {"stateMachineArn": state_machine_arn, "statusFilter": "FAILED", "maxResults": 100}
        if token:
            kwargs["nextToken"] = token
        resp = sfn.list_executions(**kwargs)
        older_than_window = False
        for ex in resp.get("executions", []):
            stat["scanned"] += 1
            stop = ex.get("stopDate") or ex.get("startDate")
            if stop is None:
                continue
            stop = stop.astimezone(timezone.utc)
            if stop < lo:
                older_than_window = True   # ここから先はさらに古い（降順）ので打ち切り
                break
            if stop <= hi:
                candidates.append((abs((stop - center_utc).total_seconds()), stop, ex))
        token = resp.get("nextToken")
        if older_than_window or not token:
            break

    if not candidates:
        return None, None, stat   # 窓内に失敗実行なし

    # 対象時刻に一番近い失敗実行を採用する
    candidates.sort(key=lambda c: c[0])
    _, stop, ex = candidates[0]
    exec_arn = ex["executionArn"]
    exec_name = ex.get("name", "")
    stat["execution"] = exec_name
    stat["stopDate"] = fmt_ts(ms(stop))

    error, cause, failed_state = _sf_failure_detail(sfn, exec_arn)
    lines = [
        "Step Functions 失敗実行",
        f"  実行名     : {exec_name}",
        f"  停止時刻   : {fmt_ts(ms(stop))} UTC",
        f"  失敗ステート: {failed_state or '(不明)'}",
        f"  error      : {error or '(なし)'}",
        f"  cause      : {cause or '(なし)'}",
    ]
    return "\n".join(lines), sf_console_link(region, exec_arn), stat


def main():
    """ツール全体の流れ。上から順に:
    準備 → 入力の確認 → 出力先の決定 → 対象行の抽出 → 1行ずつログ取得・転記 → 保存・集計。
    """
    log("=" * 70)
    log(" スプレッドシート → CloudWatch ログ転記ツール")
    log("=" * 70)
    setup_ssl()

    # 必要なライブラリを読み込む（未インストールなら分かりやすく案内して終了）
    try:
        import boto3                                   # AWSと通信する
        from openpyxl import load_workbook             # Excel(.xlsx)を読み書きする
        from openpyxl.styles import Alignment, Font    # セルの折り返し表示・文字色
    except ImportError:
        log("[NG  ] パッケージ未導入: pip install -r requirements.txt")
        sys.exit(1)

    if not INPUT_XLSX or not os.path.exists(INPUT_XLSX):
        log(f"[NG  ] 入力xlsxが見つかりません: {INPUT_XLSX!r}")
        log("       settings.py の INPUT_XLSX に、ダウンロードした .xlsx のパスを設定してください。")
        sys.exit(1)

    # 出力先の決定（OUTPUT_DIR があればそのフォルダへ。無ければ入力と同じ場所）
    if OUTPUT_DIR:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        out_name = os.path.basename(OUTPUT_XLSX) if OUTPUT_XLSX \
            else (os.path.splitext(os.path.basename(INPUT_XLSX))[0] + "_転記済.xlsx")
        out_path = os.path.join(OUTPUT_DIR, out_name)
    else:
        out_path = OUTPUT_XLSX or (os.path.splitext(INPUT_XLSX)[0] + "_転記済.xlsx")
    # 出力先が入力ファイルと同じだと入力を上書きしかねないので防ぐ
    if os.path.abspath(out_path) == os.path.abspath(INPUT_XLSX):
        log("[NG  ] 出力先が入力ファイルと同じです。OUTPUT_DIR / OUTPUT_XLSX を別の場所・名前にしてください。")
        sys.exit(1)
    # 入力ファイルは変更したくないので、コピーを作り、そのコピーにだけ書き込む
    shutil.copyfile(INPUT_XLSX, out_path)

    wb = load_workbook(out_path)  # コピー側を開く（このコピーに書き込む）
    # 対象シートを決める。SHEET_NAME が空なら、当月(例 "2026-07")のシートを自動で選ぶ
    sheet = SHEET_NAME or datetime.now(timezone.utc).astimezone().strftime("%Y-%m")
    if sheet not in wb.sheetnames:
        log(f"[NG  ] シート '{sheet}' が見つかりません。存在: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet]
    ws.column_dimensions["D"].width = 100   # D列(ログ本文)を広げて読みやすく
    ws.column_dimensions["F"].width = 12    # F列(リンク)の幅

    log("[CONF] 設定内容:")
    log(f"        入力     : {INPUT_XLSX}")
    log(f"        出力(コピー): {out_path}")
    log(f"        シート   : {sheet}   リージョン: {REGION}")
    log(f"        取得行数 : 前{LINES_BEFORE} / 後{LINES_AFTER}   検索窓: ±{WINDOW_MIN}分   MAX_ROWS: {MAX_ROWS}")
    log(f"        プロファイル: {PREFIX_TO_PROFILE}")

    # --- 処理する行を集める：A列(日時)とB列(アラーム)があり、D列(ログ)がまだ空の行 ---
    #   「D列が空 = まだ処理していない」という取り決め。処理済みの行は自動的に飛ばされる。
    targets = []
    for r in range(2, ws.max_row + 1):   # 2行目から（1行目は見出し）
        a = ws.cell(row=r, column=COL_A).value
        b = ws.cell(row=r, column=COL_B).value
        d = ws.cell(row=r, column=COL_D).value
        if a and b and (d is None or str(d).strip() == ""):
            targets.append((r, a, b))
        if MAX_ROWS and len(targets) >= MAX_ROWS:   # MAX_ROWS=0 のときは上限なし
            break
    limit_label = "制限なし" if not MAX_ROWS else str(MAX_ROWS)
    log(f"[INFO] 未処理行(D空) を {len(targets)} 行 抽出（1回の上限: {limit_label}）")
    if not targets:
        log("[INFO] 対象なし。終了します。")
        return

    # プロファイルごとのAWS接続を使い回すための入れ物
    #   （同じプロファイルは1回だけ作る。こうすると MFA も同じアカウントにつき1回で済む）
    sessions = {}

    def get_session(profile):
        """指定プロファイルのAWS接続を返す（初回だけ作成し、以降は使い回す）。"""
        if profile not in sessions:
            log(f"[AWS ] boto3セッション作成 profile={profile or '(default)'}")
            sessions[profile] = boto3.Session(profile_name=profile) if profile else boto3.Session()
        return sessions[profile]

    # 集計カテゴリ:
    #   ok    = ログを転記できた
    #   none  = 該当ログなし（Dに記録）
    #   ng    = 取得できず・要確認（未対応/アラーム未検出/時刻不正 → Dに理由を記録。再実行では変わらない）
    #   retry = 一時的なエラー（AWS通信/認証 → D空のまま。次回実行で再取得）
    counts = {"ok": 0, "none": 0, "ng": 0, "retry": 0}

    def write_reason(row, body):
        """ログを取得できなかった行に、理由をD列へ記入（人が見て対応でき、二重処理も防ぐ）。"""
        c = ws.cell(row=row, column=COL_D)
        c.value = body
        c.alignment = Alignment(wrap_text=True, vertical="top")

    def write_log(row, body, url):
        """取得できた本文を D列に、コンソールへのリンクを F列に書き込む（通常ログ・SF共通）。
        セルの文字数上限を超える本文は途中で切り、全文は F列リンクから見てもらう。"""
        if len(body) > XLSX_CELL_LIMIT:
            log(f"[WARN] 本文が長い({len(body)}字) → セル上限のため {XLSX_CELL_LIMIT}字で切詰（全文はF列リンク参照）")
            body = body[:XLSX_CELL_LIMIT] + "\n…(セル文字数上限で以降省略。全文はF列リンク参照)"
        dcell = ws.cell(row=row, column=COL_D)
        dcell.value = body
        dcell.alignment = Alignment(wrap_text=True, vertical="top")   # 改行表示で読みやすく
        fcell = ws.cell(row=row, column=COL_F)
        fcell.value = "リンク"                     # セルには「リンク」と表示
        fcell.hyperlink = url                      # クリックでコンソールを開く
        fcell.font = Font(color="0563C1", underline="single")

    # ここから1行ずつ処理する
    for (r, a, b) in targets:
        alarm = str(b).strip()
        prefix = alarm.split("-", 1)[0]                          # アラーム名の先頭（例 "strike"）
        profile = PREFIX_TO_PROFILE.get(prefix, DEFAULT_PROFILE)  # 対応するAWSプロファイル
        center = parse_utc(a)                                    # A列の日時（UTC）
        log("")
        log("-" * 70)
        log(f"[行 {r}] alarm={alarm}")
        log(f"        接頭辞={prefix} → profile={profile or '(default)'}   A(UTC)={a}")
        if center is None:
            write_reason(r, f"【取得できず】A列の日時を解釈できませんでした（値: {a}）。"
                            f"日付の形式をご確認ください。")
            log("[NG  ] A列の時刻を解釈できず → D に理由を記入（再実行では変わりません）")
            counts["ng"] += 1
            continue

        try:
            sess = get_session(profile)
            cw = sess.client("cloudwatch", region_name=REGION)
            logs = sess.client("logs", region_name=REGION)
        except Exception as e:
            log(f"[NG  ] AWSセッション失敗 → D空のまま（次回再実行で再取得）: {e!r}")
            counts["retry"] += 1
            continue

        # アラーム→ロググループ解決
        try:
            lg, fp, stype, note, detail = resolve_log_group(logs, cw, alarm)
        except Exception as e:
            kind, msg = classify_aws_error(e)
            if kind == "permanent":
                write_reason(r, f"【取得できず】{msg}")
                log(f"[NG  ] ロググループ解決失敗（{msg}）→ D に理由を記入（要確認）")
                counts["ng"] += 1
            else:
                log(f"[NG  ] ロググループ解決失敗（AWS呼び出しエラー）→ D空のまま（次回再実行で再取得）: {e!r}")
                counts["retry"] += 1
            continue
        if detail:
            log(f"[STEP] describe_alarms: namespace={detail.get('namespace')} "
                f"metric={detail.get('metric')} dimensions={detail.get('dimensions')}")

        # --- Step Functions は CloudWatch Logs ではなく「実行履歴」から失敗内容を取得する ---
        if stype == "states":
            arn = (detail.get("dimensions") or {}).get("StateMachineArn")
            if not arn:
                write_reason(r, "【取得できず】Step Functions ですが、アラームから StateMachineArn を"
                                "取得できませんでした。実行画面を手動でご確認ください。")
                log("[NG  ] Step Functions: StateMachineArn 不明 → D に理由を記入（要確認）")
                counts["ng"] += 1
                continue
            try:
                sfn = sess.client("stepfunctions", region_name=REGION)
                sf_text, sf_url, sf_stat = fetch_stepfunctions_error(sfn, REGION, arn, center)
            except Exception as e:
                kind, msg = classify_aws_error(e)
                if kind == "permanent":
                    write_reason(r, f"【取得できず】{msg}")
                    log(f"[NG  ] Step Functions 取得失敗（{msg}）→ D に理由を記入（要確認）")
                    counts["ng"] += 1
                else:
                    log(f"[NG  ] Step Functions 取得失敗（AWS呼び出しエラー）→ D空のまま（次回再実行で再取得）: {e!r}")
                    counts["retry"] += 1
                continue
            log(f"[STEP] Step Functions: 失敗実行を走査 {sf_stat.get('scanned')}件  窓={sf_stat.get('window')}")
            if sf_text is None:
                write_reason(r, f"【該当ログなし】Step Functions の失敗実行が時間窓内に見つかりませんでした"
                                f"（窓={sf_stat.get('window')}）。取得日時={datetime.now().strftime('%Y-%m-%d %H:%M')}")
                log("[WARN] Step Functions: 該当する失敗実行なし → D に『該当なし』を記入")
                counts["none"] += 1
                continue
            log(f"[OK  ] 転記: D{r}=Step Functions失敗内容(実行名={sf_stat.get('execution')}) / F{r}=「リンク」(実行画面)")
            log(f"        F{r} → {sf_url}")
            write_log(r, sf_text, sf_url)
            counts["ok"] += 1
            continue

        if not lg:
            if stype == "mf_error":
                # メトリクスフィルタ取得でAWSエラー（権限不足や一時障害の可能性）→ 再実行対象
                log(f"[NG  ] ロググループ特定不可（AWSエラー: {note}）→ D空のまま（次回再実行で再取得）")
                counts["retry"] += 1
            else:
                # アラーム未検出 / 権限 / 未対応の種類 → 再実行しても変わらないのでDに理由を記録
                if stype == "unknown":
                    reason = ("アラームが見つかりません（削除済み・名前違い、または対象アカウント/"
                              "プロファイルの不一致の可能性）。")
                elif stype == "mf_denied":
                    reason = note
                else:
                    reason = f"未対応の種類のためロググループを特定できませんでした（{note}）。"
                write_reason(r, f"【取得できず】{reason}")
                log(f"[NG  ] ロググループ特定不可（{stype}: {note}）→ D に理由を記入（要確認）")
                counts["ng"] += 1
            continue
        log(f"[STEP] ロググループ={lg}  type={stype}  filter={fp!r}")

        # ログ取得
        try:
            text, link, _n, stat = fetch_logs(logs, lg, fp, center)
        except Exception as e:
            kind, msg = classify_aws_error(e)
            if kind == "permanent":
                write_reason(r, f"【取得できず】{msg}")
                log(f"[NG  ] ログ取得失敗（{msg}）→ D に理由を記入（要確認）")
                counts["ng"] += 1
            else:
                log(f"[NG  ] ログ取得失敗（AWS呼び出しエラー）→ D空のまま（次回再実行で再取得）: {e!r}")
                counts["retry"] += 1
            continue
        log(f"[STEP] filter_log_events: 窓={stat.get('window')}  該当={stat.get('matched')}件")

        if text is None:
            write_reason(r, f"【該当ログなし】ロググループ={lg} 窓=±{WINDOW_MIN}分。"
                            f"この時間帯に一致するログが見つかりませんでした"
                            f"（ログ未出力、または検索窓が短い可能性）。"
                            f"取得日時={datetime.now().strftime('%Y-%m-%d %H:%M')}")
            log(f"[WARN] 窓内に該当イベントなし → D{r} に『該当なし』を記入（二重抽出防止）")
            counts["none"] += 1
            continue

        log(f"[STEP] get_log_events: stream={stat.get('stream')}  該当時刻={stat.get('hit_time')}  "
            f"ストリーム取得={stat.get('stream_events')}件 → 抽出={stat.get('extracted')}行")

        url = console_link(REGION, *link)
        log(f"[OK  ] 転記: D{r}=ログ本文({len(text)}字・折返し表示) / F{r}=「リンク」(ハイパーリンク)")
        log(f"        F{r} → {url}")
        write_log(r, text, url)
        counts["ok"] += 1

    try:
        wb.save(out_path)
    except PermissionError:
        log(f"[NG  ] 出力ファイルに書き込めません（開いていませんか？）: {out_path}")
        log("       Excel等で開いている場合は閉じてから、もう一度実行してください。")
        sys.exit(1)
    log("")
    log("=" * 70)
    log(f"[DONE] 転記OK={counts['ok']}  該当ログなし={counts['none']}  "
        f"取得できず・要確認={counts['ng']}  要再実行(D空)={counts['retry']}")
    log(f"[DONE] 出力ファイル: {out_path}")
    log("       ・D列に本文が入った行 → 確認のうえ、本物のシートの該当行のD/F列だけに貼り付け（丸ごと再uploadはしない）。")
    log("       ・『【取得できず】』の行 → 理由を確認し、必要なら手動対応。原因を直したらD列を空にすると次回再取得します。")
    if counts["retry"]:
        log(f"       ・要再実行が {counts['retry']} 件 → 一時的なエラーです。もう一度実行してください（D空なので再取得します）。")
    log("=" * 70)


if __name__ == "__main__":
    main()
