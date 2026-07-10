# -*- coding: utf-8 -*-
"""
検証プロトタイプ: スプレッドシート(.xlsx) の未処理行から CloudWatch ログを引っ張り、
                 入力xlsxのコピーに D(ログ)・F(リンク) を実際に転記する（v2ツールの中核）

改良点:
  - 入力xlsxを丸ごとコピーし、そのコピーの D列/F列へ実際に書き込んで「転記」を検証
  - ログを大幅に充実（設定内容・アラーム解決・ロググループ・イベント数・書込セル・集計）
AWS認証は各自の環境（プロファイル/スイッチロール/MFA）を利用。
※ AWSの資格情報はこのスクリプトからは読み書きしません。boto3が実行時に ~/.aws を参照します。
※ MFAはロール引き受けのたびに要求されます。TOTPは使い回せないため、2回目は新しいコードを入力してください。
"""

import os
import sys
import shutil
from datetime import datetime, timezone, timedelta
from urllib.parse import quote

# ===== ここだけ編集してください ============================================
INPUT_XLSX   = r"G:\マイドライブ\04_作業履歴\20260629_人間AWS自動化\検証\ログ取得検証\slack通知アラート収集_2026年度.xlsx"
SHEET_NAME   = ""            # "" なら実行時の年月(JST)から "YYYY-MM" を自動選択。明示するなら "2026-07" 等
# 出力（入力のコピーに転記したもの）。"" なら 入力名 + "_転記済.xlsx"
OUTPUT_XLSX  = ""

REGION       = "ap-northeast-1"
LINES_BEFORE = 100           # 該当行の前 何行（N）
LINES_AFTER  = 100           # 該当行の後 何行（N）
WINDOW_MIN   = 5             # A列時刻の前後 何分を検索窓にするか
MAX_ROWS     = 0             # 1回で処理する最大行数。0 = 制限なし（運用の既定）。動作確認は 3 等に

# アラーム名の接頭辞 → AWSプロファイル名（~/.aws/config のスイッチロール用プロファイル）
PREFIX_TO_PROFILE = {
    "strike": "yzk-strike-stg",
    "duel":   "yzk-duel-stg",
    "aegis":  "yzk-aegis-stg",
    "blitz":  "yzk-blitz-stg",
    "buster": "yzk-buster-stg",
}
DEFAULT_PROFILE = ""         # 上記に無い接頭辞のときのプロファイル（"" なら既定資格情報）

# 社内TLS検査プロキシ対策（AWS通信用）
CA_BUNDLE    = ""            # 社内CA(.pem)があればパス。空なら truststore(OS証明書ストア)を使用

XLSX_CELL_LIMIT = 32000      # Excelセルの文字数上限(32767)手前で切る
# ==========================================================================

COL_A, COL_B, COL_D, COL_F = 1, 2, 4, 6   # 列番号(1始まり): A=日時, B=アラーム, D=ログ本文, F=リンク


def log(msg=""):
    print(msg, flush=True)


def setup_ssl():
    if CA_BUNDLE:
        os.environ["AWS_CA_BUNDLE"] = CA_BUNDLE
        os.environ["REQUESTS_CA_BUNDLE"] = CA_BUNDLE
        log(f"[SSL ] CA_BUNDLE を使用: {CA_BUNDLE}")
    else:
        try:
            import truststore
            truststore.inject_into_ssl()
            log("[SSL ] truststore: OSの証明書ストアを使用")
        except ImportError:
            log("[SSL ] truststore無し。SSLで失敗する場合は `pip install truststore` か CA_BUNDLE指定")


def parse_utc(val):
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val.astimezone(timezone.utc)
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def ms(dt):
    return int(dt.timestamp() * 1000)


def fmt_ts(epoch_ms):
    return datetime.fromtimestamp(epoch_ms / 1000, timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def console_link(region, log_group, log_stream, start_ms, end_ms):
    def enc(s):
        return quote(quote(s, safe=""), safe="").replace("%", "$")
    base = f"https://{region}.console.aws.amazon.com/cloudwatch/home?region={region}#logsV2:log-groups"
    url = f"{base}/log-group/{enc(log_group)}/log-events/{enc(log_stream)}"
    url += f"$3FstartTime$3D{start_ms}$26endTime$3D{end_ms}"
    return url


def resolve_log_group(logs, cw, alarm_name):
    """アラーム→(log_group, filter_pattern, source_type, note, detail_dict) を返す。"""
    da = cw.describe_alarms(AlarmNames=[alarm_name])
    metric_alarms = da.get("MetricAlarms", [])
    if not metric_alarms:
        return None, None, "unknown", "アラームが見つからない/メトリクスアラームでない", {}
    a = metric_alarms[0]
    ns = a.get("Namespace")
    mn = a.get("MetricName")
    dims = {d["Name"]: d["Value"] for d in a.get("Dimensions", [])}
    detail = {"namespace": ns, "metric": mn, "dimensions": dims}

    if ns == "AWS/Lambda" and "FunctionName" in dims:
        return f"/aws/lambda/{dims['FunctionName']}", "", "lambda_native", "AWS/Lambda ネイティブ", detail
    if ns == "AWS/States":
        return None, None, "states", "Step Functions（本検証ではログ取得未対応）", detail

    # ログのメトリクスフィルタ型（カスタムNamespace）
    try:
        mf = logs.describe_metric_filters(metricName=mn, metricNamespace=ns)
        filters = mf.get("metricFilters", [])
        if filters:
            f = filters[0]
            detail["metric_filter"] = f.get("filterName")
            return f.get("logGroupName"), f.get("filterPattern", ""), "metric_filter", f"namespace={ns}", detail
    except Exception as e:
        return None, None, "mf_error", f"describe_metric_filters失敗: {e!r}", detail

    return None, None, "unsupported", f"未対応 namespace={ns} metric={mn}", detail


def fetch_logs(logs, log_group, filter_pattern, center_utc):
    """(text, link_tuple, n_matches, stat_dict) を返す。該当なしは text=None。"""
    start = center_utc - timedelta(minutes=WINDOW_MIN)
    end = center_utc + timedelta(minutes=WINDOW_MIN)
    kwargs = dict(logGroupName=log_group, startTime=ms(start), endTime=ms(end), limit=50)
    if filter_pattern:
        kwargs["filterPattern"] = filter_pattern
    fl = logs.filter_log_events(**kwargs)
    events = fl.get("events", [])
    stat = {"window": f"{fmt_ts(ms(start))}〜{fmt_ts(ms(end))} UTC", "matched": len(events)}
    if not events:
        return None, None, 0, stat

    first = events[0]
    stream = first["logStreamName"]
    hit_ts = first["timestamp"]
    stat["stream"] = stream
    stat["hit_time"] = fmt_ts(hit_ts)

    ge = logs.get_log_events(
        logGroupName=log_group, logStreamName=stream,
        startTime=hit_ts - WINDOW_MIN * 60 * 1000,
        endTime=hit_ts + WINDOW_MIN * 60 * 1000,
        startFromHead=True, limit=10000,
    )
    stream_events = ge.get("events", [])
    stat["stream_events"] = len(stream_events)
    idx = 0
    for i, e in enumerate(stream_events):
        if e["timestamp"] >= hit_ts:
            idx = i
            break
    lo = max(0, idx - LINES_BEFORE)
    hi = min(len(stream_events), idx + LINES_AFTER + 1)
    stat["extracted"] = hi - lo
    lines = []
    for i in range(lo, hi):
        e = stream_events[i]
        mark = ">>> " if i == idx else "    "
        lines.append(f"{mark}[{fmt_ts(e['timestamp'])}] {e['message'].rstrip()}")
    text = "\n".join(lines)
    link = (log_group, stream, ms(start), ms(end))
    return text, link, len(events), stat


def main():
    log("=" * 70)
    log(" スプシ → CloudWatch ログ転記 検証ツール")
    log("=" * 70)
    setup_ssl()

    try:
        import boto3
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log("[NG  ] パッケージ未導入: pip install boto3 openpyxl truststore")
        sys.exit(1)

    if not os.path.exists(INPUT_XLSX):
        log(f"[NG  ] 入力xlsxが見つかりません: {INPUT_XLSX}")
        sys.exit(1)

    out_path = OUTPUT_XLSX or (os.path.splitext(INPUT_XLSX)[0] + "_転記済.xlsx")
    # 入力を汚さないようコピーして、そのコピーに転記する
    shutil.copyfile(INPUT_XLSX, out_path)

    wb = load_workbook(out_path)  # 書き込み可能で開く（コピー側）
    sheet = SHEET_NAME or datetime.now(timezone.utc).astimezone().strftime("%Y-%m")
    if sheet not in wb.sheetnames:
        log(f"[NG  ] シート '{sheet}' が見つかりません。存在: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet]
    ws.column_dimensions["D"].width = 100   # ログ本文を読みやすく
    ws.column_dimensions["F"].width = 12

    log("[CONF] 設定内容:")
    log(f"        入力     : {INPUT_XLSX}")
    log(f"        出力(コピー): {out_path}")
    log(f"        シート   : {sheet}   リージョン: {REGION}")
    log(f"        取得行数 : 前{LINES_BEFORE} / 後{LINES_AFTER}   検索窓: ±{WINDOW_MIN}分   MAX_ROWS: {MAX_ROWS}")
    log(f"        プロファイル: {PREFIX_TO_PROFILE}")

    # --- 未処理行(D空)の抽出 ---
    targets = []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=COL_A).value
        b = ws.cell(row=r, column=COL_B).value
        d = ws.cell(row=r, column=COL_D).value
        if a and b and (d is None or str(d).strip() == ""):
            targets.append((r, a, b))
        if MAX_ROWS and len(targets) >= MAX_ROWS:
            break
    limit_label = "制限なし" if not MAX_ROWS else str(MAX_ROWS)
    log(f"[INFO] 未処理行(D空) を {len(targets)} 行 抽出（1回の上限: {limit_label}）")
    if not targets:
        log("[INFO] 対象なし。終了します。")
        return

    sessions = {}

    def get_session(profile):
        if profile not in sessions:
            log(f"[AWS ] boto3セッション作成 profile={profile or '(default)'}")
            sessions[profile] = boto3.Session(profile_name=profile) if profile else boto3.Session()
        return sessions[profile]

    # 集計カテゴリ:
    #   ok    = ログを転記できた
    #   none  = 該当ログなし（Dに記録）
    #   ng    = 取得できず・要確認（未対応/アラーム未検出/時刻不正 → Dに理由を記録。再実行では変わらない）
    #   retry = 一時的なエラー（AWS通信/認証 → D空のまま。次回実行で再取得）
    counts = {"ok": 0, "none": 0, "ng": 0, "retry": 0}

    def write_reason(row, body):
        """ログを取得できなかった行に、理由をD列へ記入（人が見て対応でき、二重処理も防ぐ）。"""
        c = ws.cell(row=row, column=COL_D)
        c.value = body
        c.alignment = Alignment(wrap_text=True, vertical="top")

    for (r, a, b) in targets:
        alarm = str(b).strip()
        prefix = alarm.split("-", 1)[0]
        profile = PREFIX_TO_PROFILE.get(prefix, DEFAULT_PROFILE)
        center = parse_utc(a)
        log("")
        log("-" * 70)
        log(f"[行 {r}] alarm={alarm}")
        log(f"        接頭辞={prefix} → profile={profile or '(default)'}   A(UTC)={a}")
        if center is None:
            write_reason(r, f"【取得できず】A列の日時を解釈できませんでした（値: {a}）。"
                            f"日付の形式をご確認ください。")
            log("[NG  ] A列の時刻を解釈できず → D に理由を記入（再実行では変わりません）")
            counts["ng"] += 1
            continue

        try:
            sess = get_session(profile)
            cw = sess.client("cloudwatch", region_name=REGION)
            logs = sess.client("logs", region_name=REGION)
        except Exception as e:
            log(f"[NG  ] AWSセッション失敗 → D空のまま（次回再実行で再取得）: {e!r}")
            counts["retry"] += 1
            continue

        # アラーム→ロググループ解決
        try:
            lg, fp, stype, note, detail = resolve_log_group(logs, cw, alarm)
        except Exception as e:
            log(f"[NG  ] ロググループ解決失敗（AWS呼び出しエラー）→ D空のまま（次回再実行で再取得）: {e!r}")
            counts["retry"] += 1
            continue
        if detail:
            log(f"[STEP] describe_alarms: namespace={detail.get('namespace')} "
                f"metric={detail.get('metric')} dimensions={detail.get('dimensions')}")
        if not lg:
            if stype == "mf_error":
                # メトリクスフィルタ取得でAWSエラー（権限不足や一時障害の可能性）→ 再実行対象
                log(f"[NG  ] ロググループ特定不可（AWSエラー: {note}）→ D空のまま（次回再実行で再取得）")
                counts["retry"] += 1
            else:
                # 未対応ソース / アラーム未検出 / 未対応namespace → 再実行しても変わらないのでDに理由を記録
                if stype == "states":
                    reason = ("このアラームは Step Functions で、本ツールはログの自動取得に未対応です。"
                              "実行画面を手動でご確認ください。")
                elif stype == "unknown":
                    reason = ("アラームが見つかりません（削除済み・名前違い、または対象アカウント/"
                              "プロファイルの不一致の可能性）。")
                else:
                    reason = f"未対応の種類のためロググループを特定できませんでした（{note}）。"
                write_reason(r, f"【取得できず】{reason}")
                log(f"[NG  ] ロググループ特定不可（{stype}: {note}）→ D に理由を記入（要確認）")
                counts["ng"] += 1
            continue
        log(f"[STEP] ロググループ={lg}  type={stype}  filter={fp!r}")

        # ログ取得
        try:
            text, link, n, stat = fetch_logs(logs, lg, fp, center)
        except Exception as e:
            log(f"[NG  ] ログ取得失敗（AWS呼び出しエラー）→ D空のまま（次回再実行で再取得）: {e!r}")
            counts["retry"] += 1
            continue
        log(f"[STEP] filter_log_events: 窓={stat.get('window')}  該当={stat.get('matched')}件")

        if text is None:
            write_reason(r, f"【該当ログなし】ロググループ={lg} 窓=±{WINDOW_MIN}分。"
                            f"この時間帯に一致するログが見つかりませんでした"
                            f"（ログ未出力、または検索窓が短い可能性）。"
                            f"取得日時={datetime.now().strftime('%Y-%m-%d %H:%M')}")
            log(f"[WARN] 窓内に該当イベントなし → D{r} に『該当なし』を記入（二重抽出防止）")
            counts["none"] += 1
            continue

        log(f"[STEP] get_log_events: stream={stat.get('stream')}  該当時刻={stat.get('hit_time')}  "
            f"ストリーム取得={stat.get('stream_events')}件 → 抽出={stat.get('extracted')}行")

        url = console_link(REGION, *link)
        if len(text) > XLSX_CELL_LIMIT:
            log(f"[WARN] ログ本文が長い({len(text)}字) → セル上限のため {XLSX_CELL_LIMIT}字で切詰（全文はF列リンク参照）")
            text = text[:XLSX_CELL_LIMIT] + "\n…(セル文字数上限で以降省略。全文はF列リンク参照)"

        dcell = ws.cell(row=r, column=COL_D)
        dcell.value = text
        dcell.alignment = Alignment(wrap_text=True, vertical="top")   # 改行表示で読みやすく
        fcell = ws.cell(row=r, column=COL_F)
        fcell.value = "リンク"                                          # セルには「リンク」と表示
        fcell.hyperlink = url                                          # クリックでコンソールを開く
        fcell.font = Font(color="0563C1", underline="single")
        log(f"[OK  ] 転記: D{r}=ログ本文({len(text)}字・折返し表示) / F{r}=「リンク」(ハイパーリンク)")
        log(f"        F{r} → {url}")
        counts["ok"] += 1

    wb.save(out_path)
    log("")
    log("=" * 70)
    log(f"[DONE] 転記OK={counts['ok']}  該当ログなし={counts['none']}  "
        f"取得できず・要確認={counts['ng']}  要再実行(D空)={counts['retry']}")
    log(f"[DONE] 出力ファイル: {out_path}")
    log("       ・D列に本文が入った行 → 確認のうえ、本物のシートの該当行のD/F列だけに貼り付け（丸ごと再uploadはしない）。")
    log("       ・『【取得できず】』の行 → 理由を確認し、必要なら手動対応。原因を直したらD列を空にすると次回再取得します。")
    if counts["retry"]:
        log(f"       ・要再実行が {counts['retry']} 件 → 一時的なエラーです。もう一度実行してください（D空なので再取得します）。")
    log("=" * 70)


if __name__ == "__main__":
    main()
